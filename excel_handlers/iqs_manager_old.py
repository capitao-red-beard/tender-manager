import warnings

warnings.filterwarnings("ignore")

import numpy as np
import numpy.core.multiarray as multiarray
import pandas as pd
import xlsxwriter
import openpyxl
import re

file_sheet_row = ''

file_name = file_sheet_row[0]
sheet_name = file_sheet_row[1]
row_number = file_sheet_row[2] - 1
ignore_list = [
    i.strip().lower().replace('\n', '').replace('€', 'eur').replace('°c', 'celsius').replace('[', '').replace(']', '')
    for i in file_sheet_row[3].split('|~|') if i != '']
extra_info_columns = [
    i.strip().lower().replace('\n', '').replace('€', 'eur').replace('°c', 'celsius').replace('[', '').replace(']', '')
    for i in file_sheet_row[4].split('|~|') if i != '']
extra_request_columns = [
    i.strip().lower().replace('\n', '').replace('€', 'eur').replace('°c', 'celsius').replace('[', '').replace(']', '')
    for i in file_sheet_row[5].split('|~|') if i != '']
payload_type = file_sheet_row[6]
transit_type = file_sheet_row[7]
trade_location = 'M:\\1 Samskip MCL\\5 Sales\\3 Marketing\\CBA\Tenders\\Tender overview general\\Tender overview\\'
config_location = '\\\\legros\\Data\\admin\\leo121\\config\\'

output_file_location = '\\\\legros\\Data\\admin\\leo121\\processed\\'

try:
    if extra_info_columns[0] == 'empty':
        eic = 'empty'
    else:
        eic = 'full'
except IndexError:
    eic = 'full'

try:
    if extra_request_columns[0] == 'empty':
        erc = 'empty'
    else:
        erc = 'full'
except IndexError:
    erc = 'full'


def read_to_dict(file_name_for_dict, sheet_name_for_dict='empty', dtype=str):
    file_to_return = ''

    if file_name_for_dict[-4:] == 'xlsx':
        tempo_df = pd.read_excel(file_name_for_dict, sheet_name=sheet_name_for_dict)
    elif file_name_for_dict[-4:] == '.csv':
        tempo_df = pd.read_csv(file_name_for_dict)

    list_a = [str(i).lower().strip().replace('\n', '') for i in tempo_df['Original']]
    list_b = [str(i).lower().strip().replace('\n', '') for i in tempo_df['Samskip']]
    file_to_return = dict(zip(list_a, list_b))
    return file_to_return


def get_excel_column(number):
    a_b_c_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                  'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    if (number > 25) & (number < 702):
        one = int(np.floor(number / 26) - 1)
        two = int(number % 26)
        return a_b_c_list[one] + a_b_c_list[two]

    elif number >= 702:
        one = int(np.floor(number / 702) - 1)
        two = int(np.floor((number - 702 - (one * 702)) / 26))
        three = int(number % 26)
        return a_b_c_list[one] + a_b_c_list[two] + a_b_c_list[three]

    else:
        return a_b_c_list[number]


def get_columns(first_col, last_col):
    a_b_c_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                  'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    list_to_return = []

    while first_col != last_col:
        if first_col[-1] != 'Z':
            list_to_return.append(first_col)
            first_col = first_col[:-1] + a_b_c_list[a_b_c_list.index(first_col[-1]) + 1]
        else:
            list_to_return.append(first_col)
            if len(first_col) == 2:
                if first_col[-2] != 'Z':
                    first_col = first_col[:-2] + a_b_c_list[a_b_c_list.index(first_col[-2]) + 1] + 'A'
            elif (len(first_col) > 2):
                print('check the get_columns function, very big excel, function needs adjustments')
            else:
                first_col = 'AA'

    list_to_return.append(first_col)
    return list_to_return


def get_column_number(column_code):
    a_b_c_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                  'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    if len(column_code) == 1:
        location = a_b_c_list.index(column_code)
    elif len(column_code) == 2:
        location = (a_b_c_list.index(column_code[0]) + 1) * 26 + a_b_c_list.index(column_code[1])
    elif len(column_code) == 3:
        print('not programmed yet')
    else:
        print('impossible length of column')

    return location


def creating_dropdown(wb, ws):
    form = []

    validations = ws.data_validations.dataValidation
    for validation in validations:
        form.append(validation.formula1)
    form = list(set(form))

    new_dict = {}
    for code in form:
        temp = []
        if code == None:
            pass
        elif len(re.findall('\w+\d+:\w+\d+', code)) > 0:
            try:
                if len(ws[code]) > 1:
                    for cell in ws[code]:
                        temp.append(cell.value)
                else:
                    for i in ws[code][0]:
                        temp.append(i.value)

                new_dict[code] = temp
            except ValueError:
                new_dict[code] = code.replace('"', '').replace("'", "").split(',')

        else:
            try:
                def_name = list(wb.defined_names[code].destinations)
                temporary_sheet = wb[def_name[0][0]]
                code2 = def_name[0][1].replace('$', '')
                if ':' in code2:
                    start = code2.split(':')[0]
                    end = code2.split(':')[1]
                    temp.append(temporary_sheet[start].value)
                    while start != end:
                        start = start[:len([i for i in start if i.isalpha()])] + str(
                            int(start[len([i for i in start if i.isalpha()]):]) + 1)
                        temp.append(temporary_sheet[start].value)
                else:
                    temp.append(temporary_sheet[code2].value)
                new_dict[code] = temp
            except KeyError:
                pass

    return new_dict


def to_cell_dropdown(filenamee, sheetnamee, rownumberr, df_len, original_columns):
    wb = openpyxl.load_workbook(filenamee)
    ws = wb[sheetnamee]
    row_num = rownumberr + 1
    d_d_dic = creating_dropdown(wb, ws)
    cell_list = []
    form_list = []
    unique_columns = []
    list_cell_name = []

    validations = ws.data_validations.dataValidation
    for validation in validations:
        for cell in validation.cells:
            cell = cell.coord
            try:
                form_list.append(d_d_dic[validation.formula1])
                cell_list.append(cell)
                if re.findall('[A-Z]+', cell)[0] not in unique_columns:
                    unique_columns.append(re.findall('[A-Z]+', cell)[0])
                    list_cell_name.append(ws[(re.findall('[A-Z]+', cell)[0] + str(row_num))].value)
            except KeyError:
                pass

    temp_dict = dict(zip(cell_list, form_list))

    new_dict = {}

    for x, column_code in enumerate(unique_columns):
        temp_list = []
        row_list = []
        for key in list(temp_dict.keys()):
            if ':' in key:
                start = key.split(':')[0]
                end = key.split(':')[1]
                if (start[:len(column_code)] == column_code) & (
                        len(column_code) == len([i for i in start if i.isalpha() == True])):
                    if start[:len([i for i in start if i.isalpha()])] == end[:len([i for i in end if i.isalpha()])]:
                        row_list.append(start[len([i for i in start if i.isalpha()]):])
                        temp_list.append(temp_dict[key])
                        while start != end:
                            start = start[:len([i for i in start if i.isalpha()])] + str(
                                int(start[len([i for i in start if i.isalpha()]):]) + 1)
                            temp_list.append(temp_dict[key])
                            row_list.append(start[len([i for i in start if i.isalpha()]):])
                    else:
                        col_list = get_columns(start[:len([i for i in start if i.isalpha()])],
                                               end[:len([i for i in end if i.isalpha()])])
                        start_num = str(int(start[len([i for i in start if i.isalpha()]):]))
                        end_num = str(int(end[len([i for i in end if i.isalpha()]):]))
                        for column in col_list:
                            temp_list = []
                            row_list = []
                            start = column + start_num
                            end = column + end_num
                            row_list.append(start[len([i for i in start if i.isalpha()]):])
                            temp_list.append(temp_dict[key])
                            while start != end:
                                start = start[:len([i for i in start if i.isalpha()])] + str(
                                    int(start[len([i for i in start if i.isalpha()]):]) + 1)
                                temp_list.append(temp_dict[key])
                                row_list.append(start[len([i for i in start if i.isalpha()]):])
                            for cells in range(row_num + 1, row_num + df_len + 1):
                                if str(cells) not in row_list:
                                    row_list.append(str(cells))
                                    temp_list.append([''])
                            row_list = [int(i) for i in row_list]
                            temp_list = [qq for _, qq in sorted(zip(row_list, temp_list), reverse=False)]
                            loc_number = get_column_number(column)
                            new_dict[original_columns[loc_number].lower()] = temp_list
                        continue

            elif (key[:len(column_code)] == column_code) & (
                    len(column_code) == len([i for i in key if i.isalpha() == True])):
                temp_list.append(temp_dict[key])
                row_list.append(key[len([i for i in key if i.isalpha()]):])

        for cells in range(row_num + 1, row_num + df_len + 1):
            if str(cells) not in row_list:
                row_list.append(str(cells))
                temp_list.append([''])
        row_list = [int(i) for i in row_list]
        temp_list = [qq for _, qq in sorted(zip(row_list, temp_list), reverse=False)]
        new_dict[list_cell_name[x].lower()] = temp_list

    return new_dict


def to_samskip_format(tender_filename, tender_sheetname=0, start_row=0, list_to_ignore=None, extra_columns_info=None,
                      extra_columns_request=None, payload='', transit=''):
    column_switch = read_to_dict(config_location + 'matched_columns.csv', 'matched_columns')

    if list_to_ignore is None:
        list_to_ignore = ['']

    if extra_columns_info is None:
        extra_columns_info = ['']

    if extra_columns_request is None:
        extra_columns_request = ['']

    output_filename = 'NEW_PROCESSED_IQS_SHEET'

    pricing__department = pd.read_excel(trade_location + 'Trades.xlsx', sheet_name='Data').drop_duplicates('Trade')

    main_index = [i.split('.')[0] for i in
                  pd.read_excel(config_location + 'IQS.xlsx', sheet_name='IQS', header=0).columns]
    main_index = [(x, i) for x, i in enumerate(main_index)]

    lower_index = [i.split('.')[0] for i in
                   pd.read_excel(config_location + 'IQS.xlsx', sheet_name='IQS', header=1).columns]
    lower_index = [(x, i) for x, i in enumerate(lower_index)]

    new = pd.read_excel(config_location + 'IQS.xlsx', sheet_name='IQS', header=[1])
    new.columns = new.columns.str.lower()

    try:
        original = pd.read_excel(tender_filename, sheet_name=tender_sheetname, skiprows=start_row, dtype=str)
    except FileNotFoundError:
        original = pd.read_csv(tender_filename, skiprows=start_row, dtype=str)

    original.columns = [
        i.strip().replace("'", '').replace('\n', '').replace('€', 'eur').replace('°c', 'celsius').replace('[',
                                                                                                          '').replace(
            ']', '') for i in original.columns.str.lower()]
    original_column_list = list(original.columns)
    original = original.rename(columns=column_switch)
    columns_needed = ['customer lane id', 'origin country', 'origin city', 'origin zip code',
                      'destination country', 'destination city', 'destination zip code',
                      'equipment requested', 'shipments / year', 'transit time', 'equipment offered',
                      'payload (in ton)', 'currency', 'round 1 rate']

    for ignore_column in list_to_ignore:
        original[ignore_column] = 'Not in tender request'

    missing_columns = [i for i in columns_needed if i not in list(original.columns)]
    remaining_columns = [COL for COL in original if COL not in columns_needed if 'unnamed:' not in COL.lower()]

    remaining_columns_string = ""
    for i in remaining_columns:
        remaining_columns_string += i + "|~|"
    remaining_columns_string = remaining_columns_string[:-3]

    if len(missing_columns) > 0:
        templist = ['TRUE'] + missing_columns + ['|~|']
        templist.append(remaining_columns_string)
        print(templist)
    elif (eic == 'empty') or (erc == 'empty'):
        templist = ['TRUE'] + ['|~|']
        templist.append(remaining_columns_string)
        print(templist)
    else:

        if payload == 'kilos':
            try:
                original['payload (in ton)'] = [int(i) / 1000 for i in original['payload (in ton)']]
            except ValueError:
                pass

        if transit != '':
            transit_df = pd.read_excel(config_location + 'Transit_mapping.xlsx', str(transit))
            transit_dict = dict(zip(transit_df['Original'], transit_df['Samskip']))
            original['transit time'].replace(transit_dict, inplace=True)

        all_columns = columns_needed + extra_columns_info + extra_columns_request
        all_columns = [i.replace('\n', '').replace("'", '') for i in all_columns]
        new_file = original[all_columns]
        final = pd.concat([new, new_file], sort=True)
        dataframe_len = len(final)
        final['origin country'].replace(country_dict, inplace=True)
        final['destination country'].replace(country_dict, inplace=True)

        final['lane'] = final['origin country'] + ' - ' + final['destination country']
        final['no.'] = list(range(1, dataframe_len + 1))
        final.drop(['pricing', 'department'], 1, inplace=True)
        final = pd.merge(final, pricing__department[['Trade', 'Department', 'Pricing']], how='left', left_on='lane',
                         right_on='Trade')
        final.drop('Trade', 1, inplace=True)
        final.columns = final.columns.str.lower()
        final = final[(~final['origin country'].isnull()) & (~final['destination country'].isnull())]
        final = final[(final['origin country'] != 'nan') & (final['destination country'] != 'nan')]
        final.fillna('', inplace=True)
        final.replace('nan', '', inplace=True)

        # Adding new columns
        index1 = list(new.columns[:25].str.strip())
        index2 = list(new.columns[25:34].str.strip())
        index3 = list(new.columns[34:].str.strip())

        reindex_order = index1 + extra_columns_info + index2 + extra_columns_request + index3
        final = final.reindex(reindex_order, axis=1)

        writer = pd.ExcelWriter(output_file_location + output_filename + '.xlsx', engine='xlsxwriter', )
        workbook = writer.book
        final.to_excel(writer, sheet_name='IQS', index=False, header=False, startrow=2, startcol=0)
        pd.DataFrame().to_excel(writer, sheet_name='Hidden_dropdowns', index=False, header=False, startrow=0,
                                startcol=0)
        hidden_worksheet = writer.sheets['Hidden_dropdowns']
        worksheet = writer.sheets['IQS']
        worksheet.protect('tender')
        unlock = workbook.add_format({'locked': False})

        main__index1 = main_index[:25]
        main__index_info = len(extra_columns_info) * [(14, 'VOLUME')]
        main__index2 = main_index[25:34]
        main__index_request = len(extra_columns_request) * [(27, 'ROUND 1 OFFER\n')]
        main__index3 = main_index[34:]
        main_index = main__index1 + main__index_info + main__index2 + main__index_request + main__index3

        lower_index1 = lower_index[:25]
        lower_index_info = [(14, i) for i in extra_columns_info]
        lower_index2 = lower_index[25:34]
        lower_index_request = [(27, i) for i in extra_columns_request]
        lower_index3 = lower_index[34:]
        lower_index = lower_index1 + lower_index_info + lower_index2 + lower_index_request + lower_index3

        for nr, col in enumerate(lower_index2 + lower_index_info + lower_index3):
            for rownr in range(0, len(final)):
                value = final[col[1].lower().strip()].iloc[rownr]
                current_cell = get_excel_column(nr + len(lower_index1) + len(lower_index_info)) + str(rownr + 3)
                worksheet.write(current_cell, value, unlock)
        df_len = len(final)
        hidden_cell = 1
        try:
            cell_dropdown = to_cell_dropdown(tender_filename, tender_sheetname, start_row, df_len, original_column_list)
            for column in cell_dropdown.keys():
                dropdown_data = cell_dropdown[column]
                if column in all_columns:
                    excel_column = get_excel_column(final.columns.get_loc(column))
                else:
                    try:
                        excel_column = get_excel_column(final.columns.get_loc(column_switch[column]))
                    except KeyError:
                        continue
                for X_num in range(3, len(cell_dropdown[column]) + 3):
                    current_cell = excel_column + str(X_num)
                    worksheet.write(current_cell, '(Select a value)', unlock)
                    start_cell = hidden_cell
                    for dcell in range(start_cell, start_cell + len(dropdown_data[X_num - 3])):
                        hidden_cell += 1
                        hidden_worksheet.write('A' + str(dcell), str(dropdown_data[X_num - 3][dcell - start_cell]))
                    worksheet.data_validation(current_cell, {'dropdown': 'True',
                                                             'validate': 'list',
                                                             'input_title': 'Please select an option',
                                                             'source': '=Hidden_dropdowns!A' + str(
                                                                 start_cell) + ':A' + str(hidden_cell - 1)})

        except TypeError:
            pass
        previous_value = 0
        loc = -1
        for col_num, value in main_index:
            loc += 1
            color_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': color_dict.get(col_num),
                'border': 1,
                'font_color': 'white',
                'locked': True})
            if value != previous_value:
                worksheet.write(0, loc, value, color_format)
                previous_value = value
            else:
                worksheet.write(0, loc, '', color_format)

        loc = -1
        for col_num, value in lower_index:
            loc += 1
            color_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': color_dict.get(col_num),
                'border': 1,
                'font_color': 'white',
                'locked': True})
            worksheet.write(1, loc, value, color_format)

        worksheet.set_column('A:C', options={'hidden': True})
        worksheet.set_column('I:N', options={'hidden': True})
        writer.save()

        print('FALSE')


color_dict = {0: '#FF0000', 1: '#FF0000', 2: '#FF0000', 3: '#244062', 4: '#244062', 5: '#244062', 6: '#244062',
              7: '#244062', 8: '#244062', 9: '#244062', 10: '#244062', 11: '#244062', 12: '#244062', 13: '#244062',
              14: '#244062', 15: '#244062', 16: '#244062', 17: '#244062', 18: '#244062', 19: '#244062', 20: '#244062',
              21: '#244062', 22: '#244062', 23: '#244062', 24: '#244062', 25: '#1F497D', 26: '#1F497D', 27: '#4F81BD',
              28: '#4F81BD', 29: '#4F81BD', 30: '#4F81BD', 31: '#4F81BD', 32: '#4F81BD', 33: '#4F81BD', 34: '#4F81BD',
              35: '#4F81BD', 36: '#006985', 37: '#006985', 38: '#006985', 39: '#006985', 40: '#006985', 41: '#006985',
              42: '#006985', 43: '#948A54', 44: '#948A54', 45: '#948A54', 46: '#948A54', 47: '#948A54', 48: '#FF0000',
              49: '#FF0000', 50: '#FF0000', 51: '#FF0000', 52: '#244062', 53: '#244062', 54: '#244062'}
country_dict = {'Afghanistan': 'AF', 'Albanië': 'AL', 'Algerije': 'DZ', 'Andorra': 'AD', 'Angola': 'AO',
                'Antigua en Barbuda': 'AG', 'Argentinië': 'AR', 'Armenië': 'AM', 'Australië': 'AU',
                'Azerbeidzjan': 'AZ', "Bahama's": 'BS', 'Bahrein': 'BH', 'Bangladesh': 'BD', 'Barbados': 'BB',
                'Belarus': 'BY', 'Belgie': 'BE', 'Belize': 'BZ', 'Benin': 'BJ', 'Bhutan': 'BT', 'Boeroendi': 'BI',
                'Bolivia': 'BO', 'Bosnië-Herzegovina': 'BA', 'Botswana': 'BW', 'Brazilië': 'BR', 'Brunei': 'BN',
                'Bulgarije': 'BG', 'Burkina Faso': 'BF', 'Cambodja': 'KH', 'Canada': 'CA',
                'Centraal-Afrikaanse Republiek': 'CF', 'Chili': 'CL', 'China': 'CN', 'Colombia': 'CO', 'Comoren': 'KM',
                'Congo (Brazzaville)': 'CG', 'Congo (Democratische Republiek)': 'CD', 'Costa Rica': 'CR', 'Cuba': 'CU',
                'Cyprus': 'CY', 'Denemarken': 'DK', 'Djibouti': 'DJ', 'Dominica': 'DM', 'Dominicaanse Republiek': 'DO',
                'Duitsland': 'DE', 'Ecuador': 'EC', 'Egypte': 'EG', 'El Salvador': 'SV', 'Equatoriaal Guinea': 'CQ',
                'Eritrea': 'ER', 'Estland': 'EE', 'Fiji-eilanden': 'FJ', 'Filipijnen': 'PH', 'Finland': 'FI',
                'Frankrijk': 'FR', 'Gabon': 'GA', 'Gambia': 'GM', 'Georgië': 'GE', 'Ghana': 'GH', 'Grenada': 'GD',
                'Griekenland': 'EL', 'Guatemala': 'GT', 'Guinee': 'GN', 'Guinee Bissau': 'GW', 'Guyana': 'GY',
                'Haïti': 'HT', 'Honduras': 'HN', 'Hongarije': 'HU', 'Ierland': 'IE', 'India': 'IN', 'Indonesië': 'ID',
                'Irak': 'IQ', 'Iran': 'IR', 'Israël': 'IL', 'Italië': 'IT', 'Ivoorkust': 'CI', 'Jamaica': 'JM',
                'Japan': 'JP', 'Jemen': 'YE', 'Jordanië': 'JO', 'Kaapverdië': 'CV', 'Kameroen': 'CM',
                'Kazachstan': 'KZ', 'Kenia': 'KE', 'Kirgizië of Kirgizstan': 'KG', 'Kiribati': 'KI', 'Koeweit': 'KW',
                'Kroatië': 'HR', 'Laos': 'LA', 'Lesotho': 'LS', 'Letland': 'LV', 'Libanon': 'LB', 'Liberia': 'LR',
                'Libië': 'LY', 'Liechtenstein': 'LI', 'Litouwen': 'LT', 'Luxemburg': 'LU', 'Macedo': 'ië',
                'Madagaskar': 'MG', 'Malawi': 'MW', 'Maldiven': 'MV', 'Maleisië': 'MY', 'Mali': 'ML', 'Malta': 'MT',
                'Marokko': 'MA', 'Marshalleilanden': 'MH', 'Mauritanië': 'MR', 'Mauritius': 'MU', 'Mexico': 'MX',
                'Micronesia': 'FM', 'Moldavië': 'MD', 'Monaco ': 'MC', 'Mongolië': 'MN', 'Mozambique': 'MZ',
                'Myanmar': 'MM', 'Namibië': 'NA', 'Nauru': 'NR', 'Nederland': 'NL', 'Nepal': 'NP', 'Nicaragua': 'NI',
                'Nieuw-Zeeland': 'NZ', 'Niger': 'NE', 'Nigeria': 'NG', 'Noord-Korea': 'KP', 'Noorwegen': 'NO',
                'Oeganda': 'UG', 'Oekraïne': 'UA', 'Oezbekistan': 'UZ', 'Oman': 'OM', 'Oost-Timor': 'TL',
                'Oostenrijk ': 'AT', 'Pakistan': 'PK', 'Palau': 'PW', 'Panama': 'PA', 'Papoea-Nieuw-Guinea': 'PG',
                'Paraguay': 'PY', 'Peru': 'PE', 'Polen': 'PL', 'Portugal': 'PT', 'Qatar': 'QA', 'Roemenië': 'RO',
                'Rusland': 'RU', 'Rwanda': 'RW', 'Saint Kitts en Nevis': 'KN', 'Saint Lucia': 'LC',
                'Saint Vincent en de Grenadines': 'VC', 'Salomonseilanden': 'SB', 'Samoa': 'WS', 'San Marino': 'SM',
                'Sao Tomé en Principe': 'ST', 'Saoedi-Arabië': 'SA', 'Senegal': 'SN', 'Servië en Montenegro': 'YU',
                'Seychellen': 'SC', 'Sierra Leone': 'SL', 'Singapore ': 'SG', 'Slovenië': 'SI', 'Slowakije': 'SK',
                'Soedan': 'SD', 'Somalië': 'SO', 'Spanje': 'ES', 'Sri Lanka': 'LK', 'Suriname': 'SR', 'Swaziland': 'SZ',
                'Syrië': 'SY', 'Tadzjikistan': 'TJ', 'Taiwan': 'TW', 'Tanzania': 'TZ', 'Thailand': 'TH', 'Togo': 'TG',
                'Tonga': 'TO', 'trinidad en Tobago': 'TT', 'Tsjaad': 'TD', 'Tsjechië': 'CZ', 'Tunesië': 'TN',
                'Turkmenistan': 'TM', 'Turkije': 'tr', 'Tuvalu': 'TV', 'United Kingdom': 'GB', 'Uruguay': 'UY',
                'Vanuatu': 'VU', 'Vaticaanstad': 'VA', 'Venezuela': 'VE', 'Verenigde Arabische Emiraten': 'AE',
                'Verenigde Staten': 'US', 'Vietnam': 'VN', 'IJsland': 'IS', 'Zambia': 'ZM', 'Zimbabwe': 'ZW',
                'Zuid-Afrika': 'ZA', 'Zuid-Korea': 'KR', 'Zweden': 'SE', 'Zwitserland': 'CH', 'Albania': 'AL',
                'Algeria': 'DZ', 'American Samoa': 'AS', 'Anguilla': 'AI', 'Antarctica': 'AQ',
                'Antigua and Barbuda': 'AG', 'Argentina': 'AR', 'Armenia': 'AM', 'Aruba': 'AW', 'Australia': 'AU',
                'Austria': 'AT', 'Azerbaijan': 'AZ', 'Bahamas': 'BS', 'Bahrain': 'BH', 'Belgium': 'BE', 'Bermuda': 'BM',
                'Bonaire': 'BQ', 'Bosnia and Herzegovina': 'BA', 'Brazil': 'BR', 'British Indian Ocean Territory': 'IO',
                'British Virgin Islands': 'VG', 'Bulgaria': 'BG', 'Burundi': 'BI', 'Cambodia': 'KH', 'Cameroon': 'CM',
                'Cape Verde': 'CV', 'Cayman Islands': 'KY', 'Central African Republic': 'CF', 'Chad': 'TD',
                'Chile': 'CL', 'Christmas Island': 'CX', 'Cocos Islands': 'CC', 'Comoros': 'KM', 'Congo': 'CG',
                'Cook Islands': 'CK', 'Croatia': 'HR', 'Curacao': 'CW', 'Czech Republic': 'CZ',
                'Democratic Republic of the Congo': 'CD', 'Denmark': 'DK', 'Dominican Republic': 'DO',
                'East Timor': 'TL', 'Egypt': 'EG', 'Equatorial Guinea': 'GQ', 'Estonia': 'EE', 'Ethiopia': 'ET',
                'Falkland Islands': 'FK', 'Faroe Islands': 'FO', 'Fiji': 'FJ', 'France': 'FR', 'French Guinea': 'GF',
                'French Polynesia': 'PF', 'Georgia': 'GE', 'Germany': 'DE', 'Gibraltar': 'GI', 'Greece': 'GR',
                'Greenland': 'GL', 'Guadeloupe': 'GP', 'Guam': 'GU', 'Guernsey': 'GG', 'Guinea': 'GN',
                'Guinea-Bissau': 'GW', 'Haiti': 'HT', 'Hong Kong': 'HK', 'Hungary': 'HU', 'Iceland': 'IS',
                'Indonesia': 'ID', 'Iraq': 'IQ', 'Ireland': 'IE', 'Isle of Man': 'IM', 'Israel': 'IL', 'Italy': 'IT',
                'Ivory Coast': 'CI', 'Jersey': 'JE', 'Jordan': 'JO', 'Kazakhstan': 'KZ', 'Kenya': 'KE', 'Kosovo': 'KS',
                'Kuwait': 'KW', 'Kyrgyzstan': 'KG', 'Latvia': 'LV', 'Lebanon': 'LB', 'Libya': 'LY', 'Lithuania': 'LT',
                'Luxembourg': 'LU', 'Macao': 'MO', 'Macedonia': 'MK', 'Madagascar': 'MG', 'Malaysia': 'MY',
                'Maldives': 'MV', 'Marshall Islands': 'MH', 'Martinique': 'MQ', 'Mauritania': 'MR', 'Mayotte': 'YT',
                'Moldova': 'MD', 'Monaco': 'MC', 'Mongolia': 'MN', 'Montenegro': 'ME', 'Montserrat': 'MS',
                'Morocco': 'MA', 'Netherlands': 'NL', 'Netherlands Antilles': 'AN', 'New Caledonia': 'NC',
                'New Zealand': 'NZ', 'Niue': 'NU', 'North Korea': 'KP', 'Northern Ireland': 'NI',
                'Northern Mariana Islands': 'MP', 'Norway': 'NO', 'Papua New Guinea': 'PG', 'Philippines': 'PH',
                'Pitcairn': 'PN', 'Poland': 'PL', 'Puerto Rico': 'PR', 'Republic of the Congo': 'CG', 'Reunion': 'RE',
                'Romania': 'RO', 'Russia': 'RU', 'Saint Barthelemy': 'BL', 'Saint Helena': 'SH',
                'Saint Kitts and Nevis': 'KN', 'Saint Martin': 'MF', 'Saint Pierre and Miquelon': 'PM',
                'Saint Vincent and the Grenadines': 'VC', 'Sao Tome and Principe': 'ST', 'Saudi Arabia': 'SA',
                'Serbia': 'RS', 'Seychelles': 'SC', 'Singapore': 'SG', 'Sint Maarten': 'SX', 'Slovakia': 'SK',
                'Slovenia': 'SI', 'Solomon Islands': 'SB', 'Somalia': 'SO', 'South Africa': 'ZA', 'South Korea': 'KR',
                'Spain': 'ES', 'Sudan': 'SD', 'Svalbard and Jan Mayen': 'SJ', 'Sweden': 'SE', 'Switzerland': 'CH',
                'Syria': 'SY', 'Tajikistan': 'TJ', 'Tokelau': 'TK', 'Trinidad and Tobago': 'TT', 'Tunisia': 'TN',
                'Turkey': 'TR', 'Turks and Caicos Islands': 'TC', 'U.S. Virgin Islands': 'VI', 'Uganda': 'UG',
                'Ukraine': 'UA', 'United Arab Emirates': 'AE', 'United States': 'US', 'Uzbekistan': 'UZ',
                'Vatican': 'VA', 'Wallis and Futuna': 'WF', 'Western Sahara': 'EH', 'Yemen': 'YE'}

qq = to_samskip_format(file_name,
                       sheet_name,
                       row_number,
                       ignore_list,
                       extra_info_columns,
                       extra_request_columns,
                       payload_type,
                       transit_type)
